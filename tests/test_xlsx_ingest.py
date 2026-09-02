import os

from openpyxl import Workbook
from rdflib import RDF, Graph, Literal, Namespace, URIRef

from generator.xlsx_generator import generate_workbook
from generator.xlsx_ingest import extract_raw_cells, interpret_master_detail, interpret_positional
from generator.xsd_generator import load_graph

SSO = Namespace("https://purl.openfaster.org/sso/")
OFR = Namespace("https://openfaster.org/realizations/schema#")
IO = "https://purl.openfaster.org/io/IO_"


def _load_structure_and_layout():
    structure = load_graph("modules/kafe.ttl")
    structure.parse("/work/institutional-ontology/institutional-ontology.owl", format="xml")
    layout = load_graph("layouts/kafe-canonical.ttl")
    return structure, layout


def test_extracts_every_nonempty_cell_regardless_of_shape(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "AnySheet"
    ws.cell(row=1, column=1, value="Header")
    ws.cell(row=2, column=1, value="HERR")
    path = os.path.join(tmp_path, "raw.xlsx")
    wb.save(path)

    graph = extract_raw_cells(path)
    values = {
        (int(graph.value(c, SSO.rowIndex)), int(graph.value(c, SSO.columnIndex))): str(graph.value(c, SSO.literalValue))
        for c in graph.subjects(None, None)
        if (c, SSO.rowIndex, None) in graph
    }
    assert values[(1, 1)] == "Header"
    assert values[(2, 1)] == "HERR"


def test_interprets_positional_layout_into_concept_linked_values(tmp_path):
    # A minimal, self-contained flat single-sheet layout -- interpret_positional
    # is a generic, still-supported function; it no longer depends on this
    # module's own (now master-detail) kafe-canonical.ttl for its tests.
    layout = Graph()
    layout.bind("sso", SSO)
    sheet = URIRef("urn:test:sheet")
    layout.add((sheet, RDF.type, SSO.Sheet))
    layout.add((sheet, SSO.sheetName, Literal("Flat")))
    layout.add((sheet, SSO.sheetPosition, Literal(1)))
    anrede_col = URIRef("urn:test:col:anrede")
    layout.add((anrede_col, SSO.sheet, sheet))
    layout.add((anrede_col, SSO.columnIndex, Literal(1)))
    layout.add((anrede_col, SSO.dataStartRow, Literal(2)))
    layout.add((anrede_col, OFR.realizesConcept, URIRef(f"{IO}0000003")))
    vorname_col = URIRef("urn:test:col:vorname")
    layout.add((vorname_col, SSO.sheet, sheet))
    layout.add((vorname_col, SSO.columnIndex, Literal(2)))
    layout.add((vorname_col, SSO.dataStartRow, Literal(2)))
    layout.add((vorname_col, OFR.realizesConcept, URIRef(f"{IO}0000001")))

    structure = load_graph("modules/kafe.ttl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Flat"
    ws.cell(row=2, column=1, value="HERR")
    ws.cell(row=2, column=2, value="Hans")
    path = os.path.join(tmp_path, "filled.xlsx")
    wb.save(path)

    raw = extract_raw_cells(path)
    obs_graph = interpret_positional(raw, layout, str(sheet), structure)

    record = URIRef("urn:record:2")
    values = {
        str(obs_graph.value(obs, OFR.observedConcept)): obs_graph.value(obs, OFR.hasValue)
        for obs in obs_graph.subjects(RDF.type, OFR.FieldObservation)
        if obs_graph.value(obs, OFR.aboutRecord) == record
    }
    assert values[f"{IO}0000003"] == URIRef(f"{IO}0000005")  # HERR resolved to the Mr. individual
    assert str(values[f"{IO}0000001"]) == "Hans"


KAFE = "https://openfaster.org/kafe/schema#"
ANTRAEGE_SHEET = f"{KAFE}ErstAntragSheet"
ID_COLUMN = f"{KAFE}ErstAntragIdColumn"
PERSONEN_SHEET = f"{KAFE}PersonenSheet"
PARENT_KEY_COLUMN = f"{KAFE}PersonenParentIdColumn"
ROLE_COLUMN = f"{KAFE}PersonenRoleColumn"


def test_interpret_master_detail_mints_partofrecord_recordposition_hasrole(tmp_path):
    structure, layout = _load_structure_and_layout()
    wb = generate_workbook(structure, layout)

    antraege = wb["Erstattungsantraege"]
    antraege.cell(row=2, column=1, value="A1")
    antraege.cell(row=3, column=1, value="A2")

    personen = wb["Personen"]
    # Row 2: taxpayer on A1
    personen.cell(row=2, column=1, value="A1")
    personen.cell(row=2, column=2, value="STEUERPFLICHTIGE_PERSON")
    personen.cell(row=2, column=3, value="HERR")
    personen.cell(row=2, column=4, value="Hans")
    personen.cell(row=2, column=5, value="Muster")
    # Row 3: legal representative on A1
    personen.cell(row=3, column=1, value="A1")
    personen.cell(row=3, column=2, value="GESETZLICHE_VERTRETUNG")
    personen.cell(row=3, column=3, value="FRAU")
    personen.cell(row=3, column=4, value="Erika")
    personen.cell(row=3, column=5, value="Vertreter")
    # Row 4: taxpayer on A2
    personen.cell(row=4, column=1, value="A2")
    personen.cell(row=4, column=2, value="STEUERPFLICHTIGE_PERSON")
    personen.cell(row=4, column=3, value="HERR")
    personen.cell(row=4, column=4, value="Peter")
    personen.cell(row=4, column=5, value="Steuer")
    # Row 5: submission-level authorised representative (blank parent ID)
    personen.cell(row=5, column=2, value="BEVOLLMAECHTIGTE_PERSON")
    personen.cell(row=5, column=3, value="FRAU")
    personen.cell(row=5, column=4, value="Anna")
    personen.cell(row=5, column=5, value="Vollmacht")

    path = os.path.join(tmp_path, "filled.xlsx")
    wb.save(path)

    raw = extract_raw_cells(path)
    submission = URIRef("urn:record:submission:1")
    obs_graph = interpret_master_detail(
        raw, layout, ANTRAEGE_SHEET, ID_COLUMN, PERSONEN_SHEET, PARENT_KEY_COLUMN, ROLE_COLUMN, structure, submission
    )

    antrag1 = URIRef("urn:record:antrag:A1")
    antrag2 = URIRef("urn:record:antrag:A2")
    assert obs_graph.value(antrag1, OFR.partOfRecord) == submission
    assert int(obs_graph.value(antrag1, OFR.recordPosition)) == 1
    assert obs_graph.value(antrag2, OFR.partOfRecord) == submission
    assert int(obs_graph.value(antrag2, OFR.recordPosition)) == 2

    taxpayer1 = URIRef("urn:record:person:2")
    legal_rep1 = URIRef("urn:record:person:3")
    taxpayer2 = URIRef("urn:record:person:4")
    authorised_rep = URIRef("urn:record:person:5")

    assert obs_graph.value(taxpayer1, OFR.partOfRecord) == antrag1
    assert obs_graph.value(taxpayer1, OFR.hasRole) == URIRef(f"{IO}0000008")
    assert obs_graph.value(legal_rep1, OFR.partOfRecord) == antrag1
    assert obs_graph.value(legal_rep1, OFR.hasRole) == URIRef(f"{IO}0000009")
    assert obs_graph.value(taxpayer2, OFR.partOfRecord) == antrag2
    assert obs_graph.value(taxpayer2, OFR.hasRole) == URIRef(f"{IO}0000008")
    assert obs_graph.value(authorised_rep, OFR.partOfRecord) == submission
    assert obs_graph.value(authorised_rep, OFR.hasRole) == URIRef(f"{IO}0000010")

    taxpayer1_values = {
        str(obs_graph.value(obs, OFR.observedConcept)): obs_graph.value(obs, OFR.hasValue)
        for obs in obs_graph.subjects(RDF.type, OFR.FieldObservation)
        if obs_graph.value(obs, OFR.aboutRecord) == taxpayer1
    }
    assert str(taxpayer1_values[f"{IO}0000001"]) == "Hans"
    assert str(taxpayer1_values[f"{IO}0000002"]) == "Muster"
