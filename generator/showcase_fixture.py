"""The real Hans Muster / Erika Vertreter / Peter Steuer / Anna Vollmacht
fixture -- shared by tests/test_round_trip.py and
scripts/export_showcase_data.py so there's exactly one real source for
this data, not two copies that can drift apart.
"""
from __future__ import annotations

import os

from openpyxl import Workbook

from generator.xlsx_generator import generate_workbook


def build_fixture_workbook(structure, layout, out_dir: str) -> str:
    wb: Workbook = generate_workbook(structure, layout)
    antraege = wb["Erstattungsantraege"]
    antraege.cell(row=2, column=1, value="A1")
    antraege.cell(row=3, column=1, value="A2")

    personen = wb["Personen"]
    personen.cell(row=2, column=1, value="A1")
    personen.cell(row=2, column=2, value="STEUERPFLICHTIGE_PERSON")
    personen.cell(row=2, column=3, value="HERR")
    personen.cell(row=2, column=4, value="Hans")
    personen.cell(row=2, column=5, value="Muster")
    personen.cell(row=3, column=1, value="A1")
    personen.cell(row=3, column=2, value="GESETZLICHE_VERTRETUNG")
    personen.cell(row=3, column=3, value="FRAU")
    personen.cell(row=3, column=4, value="Erika")
    personen.cell(row=3, column=5, value="Vertreter")
    personen.cell(row=4, column=1, value="A2")
    personen.cell(row=4, column=2, value="STEUERPFLICHTIGE_PERSON")
    personen.cell(row=4, column=3, value="HERR")
    personen.cell(row=4, column=4, value="Peter")
    personen.cell(row=4, column=5, value="Steuer")
    personen.cell(row=5, column=2, value="BEVOLLMAECHTIGTE_PERSON")
    personen.cell(row=5, column=3, value="FRAU")
    personen.cell(row=5, column=4, value="Anna")
    personen.cell(row=5, column=5, value="Vollmacht")

    path = os.path.join(out_dir, "fixture.xlsx")
    wb.save(path)
    return path
