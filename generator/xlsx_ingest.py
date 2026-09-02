"""Raw .xlsx -> SSO-shaped RDF extraction. Mechanical and structure-agnostic:
every non-empty cell across every sheet becomes a triple set, with zero
interpretation of what any cell means. Works on any .xlsx regardless of
shape -- this is the step that makes "a sheet has no assumed structure"
concrete.
"""
from __future__ import annotations

from openpyxl import load_workbook
from rdflib import RDF, BNode, Graph, Literal, Namespace, URIRef

from generator.xsd_generator import XSDO

SSO = Namespace("https://purl.openfaster.org/sso/")
OFR = Namespace("https://openfaster.org/realizations/schema#")


def extract_raw_cells(path: str) -> Graph:
    graph = Graph()
    graph.bind("sso", SSO)
    wb = load_workbook(path, data_only=True)

    workbook_node = URIRef(f"urn:sso:workbook:{path}")

    for position, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        sheet_node = URIRef(f"{workbook_node}/sheet/{sheet_name}")
        graph.add((workbook_node, SSO.hasSheet, sheet_node))
        graph.add((sheet_node, SSO.sheetName, Literal(sheet_name)))
        graph.add((sheet_node, SSO.sheetPosition, Literal(position)))

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell_node = URIRef(f"{sheet_node}/cell/{cell.row}/{cell.column}")
                graph.add((sheet_node, SSO.hasCell, cell_node))
                graph.add((cell_node, SSO.rowIndex, Literal(cell.row)))
                graph.add((cell_node, SSO.columnIndex, Literal(cell.column)))
                graph.add((cell_node, SSO.literalValue, Literal(str(cell.value))))
                graph.add((cell_node, SSO.valueType, SSO.stringValue))

    return graph


def _resolve_enumeration_value(structure_graph: Graph, concept, raw_value: str):
    """If `concept` is realized by an ElementDeclaration whose type has
    xsdo:EnumerationValue individuals, find the one whose literalValue
    matches raw_value and return the IO: individual IT realizes. Returns
    None for free-text concepts (no enumeration to resolve against).
    """
    for element_decl in structure_graph.subjects(OFR.realizesConcept, concept):
        xsd_type = structure_graph.value(element_decl, XSDO["type"])
        if xsd_type is None:
            continue
        for enum_value in structure_graph.objects(xsd_type, XSDO.hasEnumerationValue):
            if str(structure_graph.value(enum_value, XSDO.literalValue)) == raw_value:
                return structure_graph.value(enum_value, OFR.realizesConcept)
    return None


def interpret_positional(
    raw_graph: Graph, layout_graph: Graph, sheet_iri: str, structure_graph: Graph
) -> Graph:
    """Walk a raw cell graph via a positional layout definition, producing a
    real ofr:FieldObservation graph -- one record per data row (record IRIs
    mechanically derived from row number). For an enumerated field, resolves
    the raw ingested token to the matching IO: individual via
    structure_graph; for a free-text field, emits the raw token as a
    literal directly. The sheet name in the layout must match the sheet
    name actually present in the raw (ingested) workbook -- this MVP
    doesn't attempt cross-sheet-name matching.
    """
    sheet_name = str(layout_graph.value(URIRef(sheet_iri), SSO.sheetName))
    raw_sheet = next(
        s for s in raw_graph.subjects(SSO.sheetName, Literal(sheet_name))
    )

    columns = list(layout_graph.subjects(SSO.sheet, URIRef(sheet_iri)))
    obs_graph = Graph()
    obs_graph.bind("ofr", OFR)

    for column in columns:
        concept = layout_graph.value(column, OFR.realizesConcept)
        col_index = int(layout_graph.value(column, SSO.columnIndex))
        data_start = int(layout_graph.value(column, SSO.dataStartRow))

        for cell in raw_graph.objects(raw_sheet, SSO.hasCell):
            row_idx = int(raw_graph.value(cell, SSO.rowIndex))
            cell_col = int(raw_graph.value(cell, SSO.columnIndex))
            if cell_col != col_index or row_idx < data_start:
                continue

            raw_value = str(raw_graph.value(cell, SSO.literalValue))
            record = URIRef(f"urn:record:{row_idx}")
            obs = BNode()
            obs_graph.add((obs, RDF.type, OFR.FieldObservation))
            obs_graph.add((obs, OFR.aboutRecord, record))
            obs_graph.add((obs, OFR.observedConcept, concept))

            resolved = _resolve_enumeration_value(structure_graph, concept, raw_value)
            obs_graph.add((obs, OFR.hasValue, resolved if resolved is not None else Literal(raw_value)))

    return obs_graph


PERSON_ROLE_CONCEPT = URIRef("https://purl.openfaster.org/io/IO_0000007")


def _row_indices(raw_graph: Graph, raw_sheet, data_start: int) -> list[int]:
    rows = {
        int(raw_graph.value(cell, SSO.rowIndex))
        for cell in raw_graph.objects(raw_sheet, SSO.hasCell)
        if int(raw_graph.value(cell, SSO.rowIndex)) >= data_start
    }
    return sorted(rows)


def _cell_value(raw_graph: Graph, raw_sheet, row: int, col_index: int) -> str | None:
    for cell in raw_graph.objects(raw_sheet, SSO.hasCell):
        if int(raw_graph.value(cell, SSO.rowIndex)) == row and int(raw_graph.value(cell, SSO.columnIndex)) == col_index:
            return str(raw_graph.value(cell, SSO.literalValue))
    return None


def interpret_master_detail(
    raw_graph: Graph,
    layout_graph: Graph,
    antraege_sheet_iri: str,
    id_column_iri: str,
    personen_sheet_iri: str,
    parent_key_column_iri: str,
    role_column_iri: str,
    structure_graph: Graph,
    submission_record: URIRef,
) -> Graph:
    """Master-detail counterpart to interpret_positional: walks an
    Erstattungsantraege (parent) + Personen (child) sheet pair, minting one
    sub-entity per Erstattungsantrag row (ofr:partOfRecord -> submission_record,
    ofr:recordPosition from real row order) and one per Personen row
    (ofr:partOfRecord -> the matching Erstattungsantrag sub-entity if the
    parent-key column is filled in, else submission_record directly -- the
    submission-level authorised representative's row), with ofr:hasRole
    resolved from the Role column. Every other Personen column becomes an
    ofr:FieldObservation about that row's person sub-entity, exactly like
    interpret_positional's leaf-field handling.
    """
    obs_graph = Graph()
    obs_graph.bind("ofr", OFR)

    antraege_sheet_name = str(layout_graph.value(URIRef(antraege_sheet_iri), SSO.sheetName))
    antraege_raw_sheet = next(raw_graph.subjects(SSO.sheetName, Literal(antraege_sheet_name)))
    id_col_index = int(layout_graph.value(URIRef(id_column_iri), SSO.columnIndex))
    id_data_start = int(layout_graph.value(URIRef(id_column_iri), SSO.dataStartRow))

    antrag_record_by_id: dict[str, URIRef] = {}
    for position, row in enumerate(_row_indices(raw_graph, antraege_raw_sheet, id_data_start), start=1):
        antrag_id = _cell_value(raw_graph, antraege_raw_sheet, row, id_col_index)
        if antrag_id is None:
            continue
        antrag_record = URIRef(f"urn:record:antrag:{antrag_id}")
        obs_graph.add((antrag_record, OFR.partOfRecord, submission_record))
        obs_graph.add((antrag_record, OFR.recordPosition, Literal(position)))
        antrag_record_by_id[antrag_id] = antrag_record

    personen_sheet_name = str(layout_graph.value(URIRef(personen_sheet_iri), SSO.sheetName))
    personen_raw_sheet = next(raw_graph.subjects(SSO.sheetName, Literal(personen_sheet_name)))
    parent_key_col_index = int(layout_graph.value(URIRef(parent_key_column_iri), SSO.columnIndex))
    role_col_index = int(layout_graph.value(URIRef(role_column_iri), SSO.columnIndex))
    personen_data_start = int(layout_graph.value(URIRef(parent_key_column_iri), SSO.dataStartRow))

    field_columns = [
        column
        for column in layout_graph.subjects(SSO.sheet, URIRef(personen_sheet_iri))
        if str(column) not in (parent_key_column_iri, role_column_iri)
        and layout_graph.value(column, OFR.realizesConcept) is not None
    ]

    for row in _row_indices(raw_graph, personen_raw_sheet, personen_data_start):
        person_record = URIRef(f"urn:record:person:{row}")

        parent_id = _cell_value(raw_graph, personen_raw_sheet, row, parent_key_col_index)
        parent_record = antrag_record_by_id[parent_id] if parent_id else submission_record
        obs_graph.add((person_record, OFR.partOfRecord, parent_record))

        role_token = _cell_value(raw_graph, personen_raw_sheet, row, role_col_index)
        if role_token is not None:
            role = _resolve_enumeration_value(structure_graph, PERSON_ROLE_CONCEPT, role_token)
            if role is not None:
                obs_graph.add((person_record, OFR.hasRole, role))

        for column in field_columns:
            concept = layout_graph.value(column, OFR.realizesConcept)
            col_index = int(layout_graph.value(column, SSO.columnIndex))
            raw_value = _cell_value(raw_graph, personen_raw_sheet, row, col_index)
            if raw_value is None:
                continue

            obs = BNode()
            obs_graph.add((obs, RDF.type, OFR.FieldObservation))
            obs_graph.add((obs, OFR.aboutRecord, person_record))
            obs_graph.add((obs, OFR.observedConcept, concept))
            resolved = _resolve_enumeration_value(structure_graph, concept, raw_value)
            obs_graph.add((obs, OFR.hasValue, resolved if resolved is not None else Literal(raw_value)))

    return obs_graph
