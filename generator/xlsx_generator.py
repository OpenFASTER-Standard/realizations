"""Graph -> canonical .xlsx generator. Reads a structural graph (XSDO-shaped,
e.g. kafe.ttl) and a layout definition (SSO-shaped, e.g.
layouts/kafe-canonical.ttl) and writes a real openpyxl Workbook: one sheet
per sso:Sheet the layout defines, headers from each DataColumn's realized
concept (or an explicit rdfs:label for structural, non-XSD columns like a
master-detail join key), and a data-validation dropdown for any column
whose realized ElementDeclaration's type has xsdo:hasEnumerationValue
individuals.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from rdflib import RDF, Graph, Namespace, URIRef

XSDO = Namespace("https://purl.openfaster.org/xsdo/")
SSO = Namespace("https://purl.openfaster.org/sso/")
OFR = Namespace("https://openfaster.org/realizations/schema#")
RDFS_LABEL = URIRef("http://www.w3.org/2000/01/rdf-schema#label")


def _concept_label(graph: Graph, concept: URIRef) -> str:
    label = graph.value(concept, RDFS_LABEL)
    return str(label) if label is not None else str(concept)


def _column_header(graph: Graph, column: URIRef, concept) -> str:
    if concept is not None:
        return _concept_label(graph, concept)
    label = graph.value(column, RDFS_LABEL)
    return str(label) if label is not None else str(column)


def _element_declaration_for_concept(graph: Graph, concept: URIRef) -> URIRef | None:
    for s in graph.subjects(OFR.realizesConcept, concept):
        if (s, RDF.type, XSDO.ElementDeclaration) in graph:
            return s
    return None


def _enumeration_tokens(graph: Graph, element_decl: URIRef) -> list[str]:
    xsd_type = graph.value(element_decl, XSDO["type"])
    if xsd_type is None:
        return []
    values = list(graph.objects(xsd_type, XSDO.hasEnumerationValue))
    tokens = [str(graph.value(v, XSDO.literalValue)) for v in values]
    return sorted(tokens)


def generate_workbook(structure_graph: Graph, layout_graph: Graph) -> Workbook:
    combined = structure_graph + layout_graph

    wb = Workbook()
    wb.remove(wb.active)

    sheets = list(layout_graph.subjects(RDF.type, SSO.Sheet))
    sheets.sort(key=lambda s: int(layout_graph.value(s, SSO.sheetPosition)))

    for sheet_uri in sheets:
        sheet_name = str(layout_graph.value(sheet_uri, SSO.sheetName))
        ws = wb.create_sheet(title=sheet_name)

        columns = list(combined.subjects(SSO.sheet, sheet_uri))
        columns.sort(key=lambda c: int(combined.value(c, SSO.columnIndex)))

        for column in columns:
            col_index = int(combined.value(column, SSO.columnIndex))
            concept = combined.value(column, OFR.realizesConcept)
            ws.cell(row=1, column=col_index, value=_column_header(combined, column, concept))

            if concept is None:
                continue
            element_decl = _element_declaration_for_concept(structure_graph, concept)
            if element_decl is None:
                continue
            tokens = _enumeration_tokens(structure_graph, element_decl)
            if not tokens:
                continue

            data_start = int(combined.value(column, SSO.dataStartRow))
            dv = DataValidation(
                type="list",
                formula1=f'"{",".join(tokens)}"',
                allow_blank=True,
            )
            col_letter = ws.cell(row=1, column=col_index).column_letter
            dv.add(f"{col_letter}{data_start}:{col_letter}1000")
            ws.add_data_validation(dv)

    return wb
